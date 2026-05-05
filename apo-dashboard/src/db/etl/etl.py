"""
APO Dashboard — Agent ETL quotidien
Lit les fichiers Excel Dropbox → détecte les modifications → met à jour Supabase

Stratégie : DELETE + INSERT par période (plus fiable que upsert sur colonnes composites)

Usage :
  python etl.py              # importe seulement les fichiers modifiés
  python etl.py --force      # réimporte tout sans vérifier les dates
"""

import os, json, hashlib, argparse, subprocess
from datetime import datetime
from pathlib import Path
import openpyxl
from supabase import create_client
from dotenv import load_dotenv

def _notify(message: str):
    try:
        subprocess.run([
            "osascript", "-e",
            f'display notification "{message}" with title "APO Dashboard"'
        ], timeout=5, capture_output=True)
    except Exception:
        pass

# ── CONFIG ────────────────────────────────────────────────────
load_dotenv(Path(__file__).parent / ".env")

SUPABASE_URL = os.environ["SUPABASE_URL"]
SUPABASE_KEY = os.environ["SUPABASE_KEY"]
sb = create_client(SUPABASE_URL, SUPABASE_KEY)

DROPBOX_COMPTA     = Path.home() / "Dropbox/APO/Compta/2026"
DROPBOX_PRODUCTION = Path.home() / "Dropbox/APO/Rapport de Production/Rapport des production 2026"
ETAT_FILE          = Path(__file__).parent / "etat_imports.json"
ANNEE_COURANTE     = datetime.now().year   # 2026
MOIS_MAX           = datetime.now().month  # mois actuel = plafond d'import

FICHIERS = {
    "caisse_graine":   DROPBOX_COMPTA / "CAISSE GRAINES 2026.xlsx",
    "caisse_apo":      DROPBOX_COMPTA / "CAISSE APO 2026.xlsx",
    "caisse_apo2":     DROPBOX_COMPTA / "CAISSE 2 APO 2026.xlsx",
    "vente_huile":     DROPBOX_COMPTA / "VENTE D'HUILE APO SARCI 2026.xlsx",
    "vente_palmiste":  DROPBOX_COMPTA / "VENTE NOIX DE PALMISTE 2026.xlsx",
    "vente_florentin": DROPBOX_COMPTA / "VENTE DE FLORENTIN 2026.xlsx",
    "vente_bassin":    DROPBOX_COMPTA / "VENTE DE BASSIN DE LAGUNAGE.xlsx",
    "pepiniere":       DROPBOX_COMPTA / "CLIENTS PEPINIERE PALMIER A HUILE.xlsx",
    "production":      DROPBOX_PRODUCTION / "Tableau de production APO 2026.xlsx",
}

# Noms des onglets Excel par source et par numéro de mois
# Les sources à onglet unique (caisse_apo, caisse_apo2, vente_huile) sont filtrées par date dans le parseur
NOMS_MOIS_FR = {
    1: 'JANVIER', 2: 'FEVRIER', 3: 'MARS',   4: 'AVRIL',
    5: 'MAI',     6: 'JUIN',    7: 'JUILLET', 8: 'AOUT',
    9: 'SEPTEMBRE', 10: 'OCTOBRE', 11: 'NOVEMBRE', 12: 'DECEMBRE',
}
LIBELLES_FR = {
    1: 'janvier', 2: 'février', 3: 'mars',   4: 'avril',
    5: 'mai',     6: 'juin',    7: 'juillet', 8: 'août',
    9: 'septembre', 10: 'octobre', 11: 'novembre', 12: 'décembre',
}

# Sources à onglet unique : même sheet pour tous les mois (filtrage par date dans le parseur)
SHEETS_PAR_MOIS = {
    "caisse_apo":  {m: "CAISSE APO"          for m in range(1, 13)},
    "caisse_apo2": {m: "CAISSE 2 APO"        for m in range(1, 13)},
    "vente_huile": {m: "VENTE D'HUILE 2026"  for m in range(1, 13)},
}

def sheet_caisse_graine(mois):
    return f"CAISSE GRAINE {NOMS_MOIS_FR[mois]}"

def sheet_production(mois):
    return f"{NOMS_MOIS_FR[mois]} 2026"

def sheet_vente_palmiste(mois):
    # Typo connue en février
    if mois == 2: return "VEMTE NOIX DE PALMISTE FEVRIER"
    return f"VENTE NOIX DE PALMISTE {NOMS_MOIS_FR[mois]}" if mois > 1 else "VENTE NOIX PALMISTE JANVIER"

def sheet_vente_florentin(mois):
    return f"VENTE FLORENTIN {NOMS_MOIS_FR[mois]}"

def sheet_vente_bassin(mois):
    return f"VENTE BASSIN {NOMS_MOIS_FR[mois]}"

# ── UTILITAIRES ───────────────────────────────────────────────

def md5_fichier(path: Path) -> str:
    h = hashlib.md5()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()

def charger_etat() -> dict:
    if ETAT_FILE.exists():
        return json.loads(ETAT_FILE.read_text())
    return {}

def sauver_etat(etat: dict):
    ETAT_FILE.write_text(json.dumps(etat, indent=2, default=str))

def get_periode_id(annee: int, mois: int) -> int | None:
    r = sb.table("periodes").select("id").eq("annee", annee).eq("mois", mois).single().execute()
    return r.data["id"] if r.data else None

def log(msg: str):
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}")

def lire_sheet(path: Path, sheet_name: str):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        log(f"  ⚠️  Sheet '{sheet_name}' introuvable dans {path.name}")
        wb.close()
        return None
    lignes = list(wb[sheet_name].iter_rows(values_only=True))
    wb.close()
    return lignes

CHUNK_SIZE = 200  # max lignes par requête Supabase (évite les 502 sur grandes tables)

def inserer(table: str, rows: list, periode_id: int | None = None, label: str = ""):
    """Supprime les anciennes lignes du mois puis insère les nouvelles par blocs."""
    if not rows:
        return
    if periode_id is not None:
        sb.table(table).delete().eq("periode_id", periode_id).execute()
    for i in range(0, len(rows), CHUNK_SIZE):
        sb.table(table).insert(rows[i:i + CHUNK_SIZE]).execute()
    log(f"  ✅ {table} : {len(rows)} lignes insérées {label}")

def categorize_libelle(libelle: str) -> str:
    """Catégorise une ligne de caisse selon la nomenclature OHADA (aligné avec categorizeLibelle() JS)."""
    l = (libelle or "").upper()

    # 66 — Charges de personnel
    if any(k in l for k in ["SALAIRE", "PAIE DU", "PAIE DES", "PAIE JOUR",
                              "PRIME POUR", "PRIME DE PRODUCTION", "PRIME MOIS", "PRIME SUR ACHAT",
                              "AVANCE SUR SALAIRE", "JOUR FERIER", "TRAVAILLEURS TEMPORAIRE",
                              "PESONNEL JOUR", "CNPS", "CMU"]):
        return "charges_personnel"

    # 61 — Frais de transport
    if any(k in l for k in ["VEHICULE", "BULDOZER", "BULL", "PORTE-CHAR", "PORTE CHAR",
                              "BILLET AVION", "LAVAGE PICK", "LOCATION CAMION",
                              "FRAIS DE TRANSPORT", "TRANSPORT FOURNISSEUR",
                              "ACHAT DE MOTO", "AFFAIRES MARITIMES", "NIVELEUSE", "DEPOT DE RAFFE"]):
        return "frais_transport"

    # 62 — Services extérieurs (entretien, maintenance, assurances)
    if any(k in l for k in ["REPARATION", "REAPARATION", "ENTRETIEN", "REBOBINAGE",
                              "RECHARGEMENT BOUTEILLE", "DEPANNAGE", "DETRATAGE",
                              "TEFLON", "FABRICATION", "CERVEAU DE FREIN",
                              "TAMPON POUR BENNE", "DECANTEUR", "FLEXIBLE POUR CHARGEUSE",
                              "PIECES DE RECHANGE", "ASSURANCE"]):
        return "services_ext"

    # 63 — Autres services extérieurs (main-d'œuvre, sécurité, divers)
    if any(k in l for k in ["MAIN D'O", "MAIN DO", "MAIN DOEUVRE",
                              "ACOMPTE SUR MAIN", "SOLDE MAIN",
                              "REBOBINEUR", "TOURNEUR", "MACONNERIE", "VITRIER",
                              "SECURIT", "GARDIEN", "NETTOYAGE", "TRAVAUX",
                              "VISA", "JURIDIQUE", "MEDICAUX", "MISSION", "HEBERGEMENT"]):
        return "autres_services_ext"

    # 65 — Autres charges (dons, relations)
    if any(k in l for k in ["BAKCHICH", "BACKCHICH", "BAKHCHICH", "FRAIS RELATIONNEL",
                              "AIDE FINANCIERE", "DON POUR", "MOBILE MONEY",
                              "RELATIONNEL", "ASSISTANCE FUNEBRE", "TEE-SHORT", "POLOS"]):
        return "autres_charges"

    # 60 — Fournitures de l'usine et des bureaux (tout le reste des achats)
    if any(k in l for k in ["CARBURANT", "GASOIL", "GAZOIL", "ESSENCE",
                              "ELECTRICIT", "CIE", "ACHAT EAU", "EAU POUR",
                              "FOURNITURES", "ACHAT DIVERS", "ENCRE", "FILTRE A EAU",
                              "MATERIELS MECANIQUE", "MATERIEL MECANIQUE", "SOUDURE",
                              "CHAUDIERE", "ORDINATEUR", "INFORMATIQUE", "BUREAU+ARMOIRE",
                              "MATERIEL ET EQUIPEMENT", "REFRIGERATEUR", "SPLIT", "FRIGO",
                              "STARLINK", "CABLE",
                              "CIMENT", "GRAVIER", "SABLE", "FORAGE", "DALLE", "BRIQUES",
                              "CONTRE PLAQUE", "TUYAUX", "CONSTRUCTION",
                              "GARAGE ENGIN", "BUREAU ANNEXE", "CERTIFICAT FONCIER",
                              "TERRAIN", "LEGALISATION", "BIDONS"]):
        return "fournitures_usine"
    if l.startswith("ACHAT"):
        return "fournitures_usine"

    return "autres_services_ext"


def safe_float(val, default=0.0) -> float:
    try:
        return float(val) if val is not None else default
    except (TypeError, ValueError):
        return default

def safe_int(val, default=0) -> int:
    try:
        return int(val) if val is not None else default
    except (TypeError, ValueError):
        return default

# ── PARSEURS ─────────────────────────────────────────────────

def parse_production_tous_mois(path: Path, periodes: dict):
    """
    Charge tous les mois de production.
    Plusieurs lignes Excel pour la même date sont agrégées :
      - flux (régimes, huile, livraisons, palmiste, stérilisateurs) → somme
      - stocks fin de journée (restant, huile, tanks, graines, palmiste) → dernière valeur
      - taux_extraction → recalculé après agrégation
    """
    sb.table("production_journaliere").delete().gte("id", 0).execute()
    log("  🗑  production_journaliere vidée")

    total = 0
    for mois, periode_id in periodes.items():
        sheet_name = sheet_production(mois)
        if not sheet_name:
            continue
        lignes = lire_sheet(path, sheet_name)
        if not lignes:
            continue

        rows_par_date = {}
        for row in lignes[2:]:
            if not row[1] or not isinstance(row[1], datetime):
                continue
            if row[1].month != mois:
                continue
            date_str = row[1].date().isoformat()

            if date_str not in rows_par_date:
                rows_par_date[date_str] = {
                    "periode_id":             periode_id,
                    "date_production":        date_str,
                    # Flux journaliers → on somme
                    "regime_recu_kg":         safe_float(row[2]),
                    "regime_traite_kg":       safe_float(row[3]),
                    "huile_produite_kg":      safe_float(row[5]),
                    "livraison_citerne_kg":   safe_float(row[7]),
                    "nb_sterilisateurs":      safe_int(row[13]),
                    "livraison_florentin_kg": safe_float(row[14]),
                    "livraison_bassin_kg":    safe_float(row[15]),
                    "production_palmiste_kg": safe_float(row[16]),
                    "livraison_palmiste_kg":  safe_float(row[17]),
                    # Stocks fin de journée → dernière valeur lue
                    "regime_restant_kg":      safe_float(row[4]),
                    "stock_huile_kg":         safe_float(row[8]),
                    "tank_1000_kg":           safe_float(row[9]),
                    "tank_300_kg":            safe_float(row[10]),
                    "stock_graine_1_kg":      safe_float(row[11]),
                    "stock_graine_2_kg":      safe_float(row[12]),
                    "stock_palmiste_kg":      safe_float(row[18]) if len(row) > 18 else 0.0,
                    # Taux d'extraction recalculé après agrégation
                    "taux_extraction":        0.0,
                }
            else:
                r = rows_par_date[date_str]
                # Flux → somme
                r["regime_recu_kg"]         += safe_float(row[2])
                r["regime_traite_kg"]       += safe_float(row[3])
                r["huile_produite_kg"]      += safe_float(row[5])
                r["livraison_citerne_kg"]   += safe_float(row[7])
                r["nb_sterilisateurs"]      += safe_int(row[13])
                r["livraison_florentin_kg"] += safe_float(row[14])
                r["livraison_bassin_kg"]    += safe_float(row[15])
                r["production_palmiste_kg"] += safe_float(row[16])
                r["livraison_palmiste_kg"]  += safe_float(row[17])
                # Stocks → dernière valeur (fin de journée)
                r["regime_restant_kg"]       = safe_float(row[4])
                r["stock_huile_kg"]          = safe_float(row[8])
                r["tank_1000_kg"]            = safe_float(row[9])
                r["tank_300_kg"]             = safe_float(row[10])
                r["stock_graine_1_kg"]       = safe_float(row[11])
                r["stock_graine_2_kg"]       = safe_float(row[12])
                if len(row) > 18:
                    r["stock_palmiste_kg"]   = safe_float(row[18])

        # Recalcule taux_extraction après agrégation complète de la journée
        for r in rows_par_date.values():
            traite = r["regime_traite_kg"]
            r["taux_extraction"] = round(r["huile_produite_kg"] / traite, 4) if traite else 0.0

        rows = list(rows_par_date.values())
        if rows:
            for i in range(0, len(rows), CHUNK_SIZE):
                sb.table("production_journaliere").insert(rows[i:i + CHUNK_SIZE]).execute()
            log(f"  ✅ production_journaliere : {len(rows)} lignes insérées (mois {mois})")
            total += len(rows)

    return total


def parse_ventes_huile(path: Path, mois: int, periode_id: int):
    sheet_name = SHEETS_PAR_MOIS["vente_huile"].get(mois)
    lignes = lire_sheet(path, sheet_name)
    if not lignes:
        return

    sarci = sb.table("clients").select("id").eq("reference", "SARCI").single().execute()
    client_id = sarci.data["id"] if sarci.data else None

    rows = []
    for row in lignes:
        if not row[1] or not isinstance(row[1], datetime):
            continue
        # Certaines lignes ont l'année 2025 par erreur de saisie (ex: janvier 2026 saisi en 2025)
        # On filtre uniquement par mois, pas par année — le sheet "2026" ne contient que des données 2026
        if row[1].month != mois:
            continue
        if safe_float(row[3]) == 0:  # ignore lignes vides
            continue
        rows.append({
            "periode_id":     periode_id,
            "client_id":      client_id,
            "date_vente":     row[1].replace(year=2026).date().isoformat(),
            "libelle":        str(row[2] or ""),
            "poids_apo_kg":   safe_float(row[3]),
            "poids_sarci_kg": safe_float(row[4]),
            "prix_kg":        safe_float(row[6]),
            "avance_sarci":   safe_float(row[7]),
        })

    inserer("ventes_huile", rows, periode_id, f"(mois {mois})")


def parse_ventes_palmiste(path: Path, mois: int, periode_id: int):
    sheet_name = sheet_vente_palmiste(mois)
    if not sheet_name:
        return
    lignes = lire_sheet(path, sheet_name)
    if not lignes:
        return

    rows = []
    for row in lignes[1:]:
        if not row[1] or not isinstance(row[1], datetime):
            continue
        if safe_float(row[3]) == 0:
            continue
        rows.append({
            "periode_id": periode_id,
            "date_vente": row[1].date().isoformat(),
            "poids_kg":   safe_float(row[3]),
            "prix_kg":    safe_float(row[4]) or 60.0,
        })

    inserer("ventes_palmiste", rows, periode_id, f"(mois {mois})")


def parse_ventes_florentin(path: Path, mois: int, periode_id: int):
    sheet_name = sheet_vente_florentin(mois)
    if not sheet_name:
        return
    lignes = lire_sheet(path, sheet_name)
    if not lignes:
        return

    rows = []
    for row in lignes[1:]:
        if not row[1] or not isinstance(row[1], datetime):
            continue
        if safe_float(row[3]) == 0:
            continue
        rows.append({
            "periode_id": periode_id,
            "date_vente": row[1].date().isoformat(),
            "poids_kg":   safe_float(row[3]),
            "prix_kg":    safe_float(row[4]),
        })

    inserer("ventes_florentin", rows, periode_id, f"(mois {mois})")


def parse_ventes_bassin(path: Path, mois: int, periode_id: int):
    sheet_name = sheet_vente_bassin(mois)
    if not sheet_name:
        return
    lignes = lire_sheet(path, sheet_name)
    if not lignes:
        return

    rows = []
    for row in lignes[1:]:
        if not row[1] or not isinstance(row[1], datetime):
            continue
        if safe_float(row[3]) == 0:
            continue
        rows.append({
            "periode_id": periode_id,
            "date_vente": row[1].date().isoformat(),
            "poids_kg":   safe_float(row[3]),
            "prix_kg":    safe_float(row[4]) or None,
            "facture":    safe_float(row[4]) > 0,
            "montant_fcfa": safe_float(row[5]),
        })

    inserer("ventes_bassin", rows, periode_id, f"(mois {mois})")


def parse_caisse_apo(path: Path, mois: int, periode_id: int):
    sheet_name = SHEETS_PAR_MOIS["caisse_apo"].get(mois)
    lignes = lire_sheet(path, sheet_name)
    if not lignes:
        return

    rows = []
    for row in lignes:
        if not row[1] or not isinstance(row[1], datetime):
            continue
        if row[1].month != mois:
            continue
        libelle = str(row[2] or "")
        rows.append({
            "periode_id":     periode_id,
            "date_mouvement": row[1].date().isoformat(),
            "libelle":        libelle,
            "debit_fcfa":     safe_float(row[3]),
            "credit_fcfa":    safe_float(row[4]),
            "solde_fcfa":     safe_float(row[5]),
            "categorie":      categorize_libelle(libelle),
        })

    inserer("caisse_apo", rows, periode_id, f"(mois {mois})")


def parse_caisse_apo2(path: Path, mois: int, periode_id: int):
    sheet_name = SHEETS_PAR_MOIS["caisse_apo2"].get(mois)
    lignes = lire_sheet(path, sheet_name)
    if not lignes:
        return

    rows = []
    for row in lignes:
        if not row[1] or not isinstance(row[1], datetime):
            continue
        if row[1].month != mois:
            continue
        libelle = str(row[2] or "")
        rows.append({
            "periode_id":     periode_id,
            "date_mouvement": row[1].date().isoformat(),
            "libelle":        libelle,
            "debit_fcfa":     safe_float(row[3]),
            "credit_fcfa":    safe_float(row[4]),
            "solde_fcfa":     safe_float(row[5]),
            "categorie":      categorize_libelle(libelle),
        })

    inserer("caisse_apo2", rows, periode_id, f"(mois {mois})")


def parse_achats_regimes(path: Path, mois: int, periode_id: int):
    sheet_name = sheet_caisse_graine(mois)
    if not sheet_name:
        return
    lignes = lire_sheet(path, sheet_name)
    if not lignes:
        return

    # Récupère ou crée les fournisseurs d'abord
    refs_vues = set()
    for row in lignes[1:]:
        if row[3] and isinstance(row[3], datetime):
            ref = str(row[4] or "INCONNU").strip()
            refs_vues.add(ref)

    for ref in refs_vues:
        sb.table("fournisseurs").upsert(
            {"reference": ref, "nom": ref}, on_conflict="reference"
        ).execute()

    # Map reference → id
    fournisseurs_map = {}
    for ref in refs_vues:
        r = sb.table("fournisseurs").select("id").eq("reference", ref).single().execute()
        if r.data:
            fournisseurs_map[ref] = r.data["id"]

    rows = []
    for row in lignes[1:]:
        if not row[3] or not isinstance(row[3], datetime):
            continue
        if safe_float(row[5]) == 0:
            continue
        ref = str(row[4] or "INCONNU").strip()
        rows.append({
            "periode_id":     periode_id,
            "fournisseur_id": fournisseurs_map.get(ref),
            "date_achat":     row[3].date().isoformat(),
            "type_transport": str(row[1] or ""),
            "numero_camion":  str(row[2] or ""),
            "poids_kg":       safe_float(row[5]),
            "prix_kg":        safe_float(row[6]),
            "prix_transport": safe_float(row[7]),
            "appro_caisse":   safe_float(row[8]),
        })

    inserer("achats_regimes", rows, periode_id, f"(mois {mois})")


def parse_pepiniere(path: Path):
    lignes = lire_sheet(path, "Feuil1")
    if not lignes:
        return

    # Supprime tout et réinsère (table cumulative sans période)
    rows_clients = []
    rows_contrats = []

    for row in lignes[1:]:
        ordre = str(row[0] or "").strip()
        if not ordre.isdigit():
            continue
        nom = str(row[2] or "").strip()
        if not nom:
            continue

        rows_clients.append({
            "reference": nom[:50],
            "nom":       nom,
            "telephone": str(row[3] or ""),
            "type":      "pepiniere",
        })
        date_contrat = row[1].date().isoformat() if isinstance(row[1], datetime) else None
        if not date_contrat:
            continue  # ignore les lignes sans date
        rows_contrats.append({
            "numero_ordre":        ordre,
            "date_contrat":        date_contrat,
            "localite_champ":      str(row[4] or ""),
            "superficie_demandee": safe_float(row[5]),
            "superficie_champ_ha": safe_float(row[7]),
            "prix_unitaire_fcfa":  safe_float(row[6]),
            "montant_total":       safe_float(row[8]),
            "net_encaisse":        safe_float(row[9]),
            "_nom_client":         nom[:50],  # clé temporaire pour jointure
        })

    # Upsert clients
    for rc in rows_clients:
        sb.table("clients").upsert(rc, on_conflict="reference").execute()

    # Récupère les ids clients et insère contrats
    sb.table("contrats_pepiniere").delete().neq("id", 0).execute()  # vide la table
    for rc in rows_contrats:
        nom_ref = rc.pop("_nom_client")
        cl = sb.table("clients").select("id").eq("reference", nom_ref).single().execute()
        rc["client_id"] = cl.data["id"] if cl.data else None
        sb.table("contrats_pepiniere").insert(rc).execute()

    log(f"  ✅ contrats_pepiniere : {len(rows_contrats)} lignes insérées")


def recalculer_kpis(mois: int, periode_id: int):
    """Recalcule kpis_mensuels depuis les tables de détail."""
    vh  = sb.table("ventes_huile").select("montant_fcfa").eq("periode_id", periode_id).execute()
    vp  = sb.table("ventes_palmiste").select("montant_fcfa").eq("periode_id", periode_id).execute()
    vf  = sb.table("ventes_florentin").select("montant_fcfa").eq("periode_id", periode_id).execute()
    vb  = sb.table("ventes_bassin").select("montant_fcfa").eq("periode_id", periode_id).execute()
    ca1 = sb.table("caisse_apo").select("credit_fcfa,libelle").eq("periode_id", periode_id).execute()
    ca2 = sb.table("caisse_apo2").select("credit_fcfa,libelle").eq("periode_id", periode_id).execute()
    bq  = sb.table("banque_apo").select("montant_fcfa,categorie").eq("periode_id", periode_id).execute()
    pj  = sb.table("production_journaliere").select("*").eq("periode_id", periode_id).execute()
    ar  = sb.table("achats_regimes").select("poids_kg,prix_kg,prix_transport").eq("periode_id", periode_id).execute()
    pep = sb.table("contrats_pepiniere").select("montant_total,net_encaisse").execute()

    # Patterns de transferts inter-caisses à exclure des charges
    SKIP = ("TRANSFERT", "VIREMENT", "VERSEMENT", "DEPOT", "APPRO")

    def is_transfer(libelle: str) -> bool:
        u = (libelle or "").upper()
        return any(u.startswith(k) for k in SKIP)

    ca_huile     = sum(safe_float(r["montant_fcfa"]) for r in vh.data)
    ca_palmiste  = sum(safe_float(r["montant_fcfa"]) for r in vp.data)
    ca_florentin = sum(safe_float(r["montant_fcfa"]) for r in vf.data)
    ca_bassin    = sum(safe_float(r["montant_fcfa"]) for r in vb.data)

    # Charges caisse (hors transferts inter-caisses)
    charges_caisse = (
        sum(safe_float(r["credit_fcfa"]) for r in ca1.data if not is_transfer(r["libelle"]))
      + sum(safe_float(r["credit_fcfa"]) for r in ca2.data if not is_transfer(r["libelle"]))
    )

    # Charges banque — séparées par type
    CATS_FIN = {"amortissement", "frais_bancaires"}
    charges_banque = sum(safe_float(r["montant_fcfa"]) for r in (bq.data or [])
                         if r.get("categorie") not in CATS_FIN)
    amort_banque   = sum(safe_float(r["montant_fcfa"]) for r in (bq.data or [])
                         if r.get("categorie") == "amortissement")
    frais_fin      = sum(safe_float(r["montant_fcfa"]) for r in (bq.data or [])
                         if r.get("categorie") == "frais_bancaires")

    # Totaux
    charges      = charges_caisse + charges_banque
    amort_total  = amort_banque or 0
    frais_total  = frais_fin

    reg_recus   = sum(safe_float(r["regime_recu_kg"])        for r in pj.data)
    reg_traites = sum(safe_float(r["regime_traite_kg"])       for r in pj.data)
    huile_prod  = sum(safe_float(r["huile_produite_kg"])      for r in pj.data)
    huile_vend  = sum(safe_float(r["livraison_citerne_kg"])   for r in pj.data)
    palm_prod   = sum(safe_float(r["production_palmiste_kg"]) for r in pj.data)
    palm_vend   = sum(safe_float(r["livraison_palmiste_kg"])  for r in pj.data)
    nb_steril   = sum(safe_int(r["nb_sterilisateurs"])        for r in pj.data)
    stock_fin   = safe_float(pj.data[-1]["regime_restant_kg"]) if pj.data else 0

    total_poids = sum(safe_float(r["poids_kg"]) for r in ar.data)
    # Coût réel = poids × (prix_kg + prix_transport) — le transport des régimes fait partie du coût MP
    total_mont  = sum(safe_float(r["poids_kg"]) * (safe_float(r["prix_kg"]) + safe_float(r["prix_transport"])) for r in ar.data)
    prix_moy    = total_mont / total_poids if total_poids else 0
    nb_camions  = len(ar.data)

    prix_huile  = ca_huile / huile_vend if huile_vend else 0
    te          = huile_prod / reg_traites if reg_traites else 0
    # Coût MP = graines nécessaires pour produire l'huile effectivement vendue
    # graine_vendu = huile_vendue / TE  →  cout_mp = graine_vendu × prix_moy
    graine_vendu = huile_vend / te if te else 0
    cout_mp      = graine_vendu * prix_moy
    ca_total    = ca_huile + ca_palmiste + ca_florentin + ca_bassin
    resultat    = ca_total - cout_mp - charges - amort_total - frais_total
    marge       = round(resultat / ca_total * 100, 2) if ca_total else 0

    pep_total    = sum(safe_float(r["montant_total"]) for r in pep.data)
    pep_encaisse = sum(safe_float(r["net_encaisse"])  for r in pep.data)

    payload = {
        "periode_id":              periode_id,
        "ca_huile_fcfa":           ca_huile,
        "ca_palmiste_fcfa":        ca_palmiste,
        "ca_florentin_fcfa":       ca_florentin,
        "ca_bassin_fcfa":          ca_bassin,
        "cout_mp_fcfa":            cout_mp,
        "charges_exploitation":    charges,
        "amortissement_fcfa":      amort_total + frais_total,
        "marge_nette_pct":         marge,
        "regimes_recus_kg":        reg_recus,
        "regimes_traites_kg":      reg_traites,
        "stock_fin_mois_kg":       stock_fin,
        "huile_produite_kg":       huile_prod,
        "huile_vendue_kg":         huile_vend,
        "taux_extraction":         round(te, 4),
        "nb_camions":              nb_camions,
        "nb_sterilisateurs":       nb_steril,
        "palmiste_produit_kg":     palm_prod,
        "palmiste_vendu_kg":       palm_vend,
        "prix_moyen_regime_kg":    round(prix_moy, 2),
        "prix_moyen_huile_kg":     round(prix_huile, 2),
        "pepiniere_contrats_fcfa": pep_total,
        "pepiniere_encaisse_fcfa": pep_encaisse,
        "pepiniere_restant_fcfa":  pep_total - pep_encaisse,
        "mis_a_jour_le":           datetime.now().isoformat(),
    }

    sb.table("kpis_mensuels").upsert(payload, on_conflict="periode_id").execute()
    log(f"  ✅ kpis_mensuels recalculés (mois {mois}) — CA : {ca_total:,.0f} | Charges : {charges:,.0f} | Amort : {amort_total:,.0f} | Résultat : {resultat:,.0f} FCFA")


# ── ORCHESTRATEUR PRINCIPAL ───────────────────────────────────

def assurer_periodes() -> dict:
    """
    Crée dans Supabase toutes les périodes de janvier jusqu'au mois courant.
    Retourne un dict {mois: periode_id} pour l'année courante.
    """
    periodes = {}
    for mois in range(1, MOIS_MAX + 1):
        # Vérifie si la période existe déjà
        pid = get_periode_id(ANNEE_COURANTE, mois)
        if not pid:
            log(f"  📅 Création période {LIBELLES_FR[mois]} {ANNEE_COURANTE}...")
            sb.table("periodes").insert({
                "annee":   ANNEE_COURANTE,
                "mois":    mois,
                "libelle": LIBELLES_FR[mois],
            }).execute()
            pid = get_periode_id(ANNEE_COURANTE, mois)
        periodes[mois] = pid
    return periodes


def run(force: bool = False):
    log("=" * 50)
    log(f"APO ETL démarré — plafond : {LIBELLES_FR[MOIS_MAX]} {ANNEE_COURANTE}")
    log("=" * 50)

    # 1. S'assure que toutes les périodes jusqu'au mois actuel existent
    periodes = assurer_periodes()
    log(f"  Périodes actives : {list(periodes.keys())}")

    etat = charger_etat()
    fichiers_modifies = []

    for cle, path in FICHIERS.items():
        if not path.exists():
            log(f"⚠️  Fichier introuvable : {path.name}")
            continue

        md5 = md5_fichier(path)
        ancien_md5 = etat.get(cle, {}).get("md5")

        if not force and md5 == ancien_md5:
            log(f"⏭  {path.name} — inchangé")
            continue

        log(f"🔄 {path.name} — modifié, import en cours...")
        fichiers_modifies.append(cle)

        # Production : traite tous les mois d'un coup (contrainte UNIQUE globale sur date)
        if cle == "production":
            try:
                parse_production_tous_mois(path, periodes)
            except Exception as e:
                log(f"  ❌ Erreur production : {e}")
            etat[cle] = {"md5": md5, "date": datetime.now().isoformat(), "fichier": path.name}
            continue

        # Pépinière : pas de boucle par mois
        if cle == "pepiniere":
            try:
                parse_pepiniere(path)
            except Exception as e:
                log(f"  ❌ Erreur pépinière : {e}")
            etat[cle] = {"md5": md5, "date": datetime.now().isoformat(), "fichier": path.name}
            continue

        # Toutes les autres sources : boucle sur les mois jusqu'au mois courant
        for mois, periode_id in periodes.items():
            try:
                if cle == "vente_huile":
                    parse_ventes_huile(path, mois, periode_id)
                elif cle == "vente_palmiste":
                    parse_ventes_palmiste(path, mois, periode_id)
                elif cle == "vente_florentin":
                    parse_ventes_florentin(path, mois, periode_id)
                elif cle == "vente_bassin":
                    parse_ventes_bassin(path, mois, periode_id)
                elif cle == "caisse_apo":
                    parse_caisse_apo(path, mois, periode_id)
                elif cle == "caisse_apo2":
                    parse_caisse_apo2(path, mois, periode_id)
                elif cle == "caisse_graine":
                    parse_achats_regimes(path, mois, periode_id)
            except Exception as e:
                log(f"  ❌ Erreur mois {mois} [{cle}] : {e}")

        etat[cle] = {"md5": md5, "date": datetime.now().isoformat(), "fichier": path.name}

    if fichiers_modifies:
        log("─" * 50)
        log("Recalcul des KPIs mensuels...")
        for mois, pid in periodes.items():
            try:
                recalculer_kpis(mois, pid)
            except Exception as e:
                log(f"  ❌ Erreur KPIs mois {mois} : {e}")
        sauver_etat(etat)
        log("=" * 50)
        log(f"ETL terminé — {len(fichiers_modifies)} fichier(s) importé(s) ✅")
        _notify(f"APO ETL ✅ — {len(fichiers_modifies)} fichier(s) importé(s)")
    else:
        log("Aucun fichier modifié — rien à faire.")
        _notify("APO ETL ✅ — Aucun fichier modifié")

    log("=" * 50)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--force", action="store_true", help="Réimporte tout sans vérifier les dates")
    args = parser.parse_args()
    run(force=args.force)
