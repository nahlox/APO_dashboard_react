"""
APO Dashboard — ETL Cloud (sans Mac)
=====================================
Télécharge les fichiers Excel depuis l'API Dropbox → parse en mémoire → Supabase.

Aucune dépendance au système local : tourne sur GitHub Actions, Railway, ou tout
autre serveur — sans que le Mac soit allumé.

Usage :
  python etl_cloud.py --tenant apo           # import complet pour le tenant 'apo'
  python etl_cloud.py --tenant apo --mois 4  # importe seulement avril
  python etl_cloud.py --tenant apo --dry     # simulation sans écriture Supabase

Prérequis (.env) :
  SUPABASE_URL     = https://xxxx.supabase.co
  SUPABASE_KEY     = eyJ...
  DROPBOX_TOKEN    = sl.xxx...  (token Dropbox App, voir README)
"""

import os, hashlib, argparse, io, time
from datetime import datetime
from pathlib import Path
import openpyxl
import dropbox
from dropbox.exceptions import ApiError
from supabase import create_client
from dotenv import load_dotenv

# ── CONFIG ────────────────────────────────────────────────────
load_dotenv(Path(__file__).parent / ".env")

SUPABASE_URL  = os.environ["SUPABASE_URL"]
SUPABASE_KEY  = os.environ["SUPABASE_KEY"]

sb = create_client(SUPABASE_URL, SUPABASE_KEY)

# Authentification Dropbox : refresh token permanent (prioritaire) ou token court
DROPBOX_REFRESH_TOKEN = os.environ.get("DROPBOX_REFRESH_TOKEN", "")
DROPBOX_APP_KEY       = os.environ.get("DROPBOX_APP_KEY", "")
DROPBOX_APP_SECRET    = os.environ.get("DROPBOX_APP_SECRET", "")
DROPBOX_TOKEN         = os.environ.get("DROPBOX_TOKEN", "")

if DROPBOX_REFRESH_TOKEN and DROPBOX_APP_KEY and DROPBOX_APP_SECRET:
    # ✅ Mode permanent — refresh token auto-renouvelé par le SDK
    dbx = dropbox.Dropbox(
        oauth2_refresh_token=DROPBOX_REFRESH_TOKEN,
        app_key=DROPBOX_APP_KEY,
        app_secret=DROPBOX_APP_SECRET,
        timeout=600,
    )
elif DROPBOX_TOKEN:
    # ⚠️ Mode temporaire — token court (expire en quelques heures)
    dbx = dropbox.Dropbox(DROPBOX_TOKEN, timeout=600)
else:
    raise EnvironmentError(
        "Aucun token Dropbox trouvé. Configure DROPBOX_REFRESH_TOKEN + DROPBOX_APP_KEY + DROPBOX_APP_SECRET "
        "dans .env (permanent) ou DROPBOX_TOKEN (temporaire)."
    )

ANNEE_COURANTE = datetime.now().year   # 2026
MOIS_MAX       = datetime.now().month  # plafond d'import = mois courant
MOIS_STATIQUES = set()                 # tous les mois importés depuis Supabase

# ── CHEMINS DROPBOX ───────────────────────────────────────────
# Chemins absolus dans ton Dropbox (sans "Dropbox/" — c'est la racine)
COMPTA_DIR     = "/APO/Compta/2026"
PRODUCTION_DIR = "/APO/Rapport de Production/Rapport des production 2026"

FICHIERS_DROPBOX = {
    "caisse_graine":   f"{COMPTA_DIR}/CAISSE GRAINES 2026.xlsx",
    "caisse_apo":      f"{COMPTA_DIR}/CAISSE APO 2026.xlsx",
    "caisse_apo2":     f"{COMPTA_DIR}/CAISSE 2 APO 2026.xlsx",
    "vente_huile":     f"{COMPTA_DIR}/VENTE D'HUILE APO SARCI 2026.xlsx",
    "vente_palmiste":  f"{COMPTA_DIR}/VENTE NOIX DE PALMISTE 2026.xlsx",
    "vente_florentin": f"{COMPTA_DIR}/VENTE DE FLORENTIN 2026.xlsx",
    "vente_bassin":    f"{COMPTA_DIR}/VENTE DE BASSIN DE LAGUNAGE.xlsx",
    "pepiniere":       f"{COMPTA_DIR}/CLIENTS PEPINIERE PALMIER A HUILE.xlsx",
    "production":      f"{PRODUCTION_DIR}/Tableau de production APO 2026.xlsx",
    "banque_apo":      f"{COMPTA_DIR}/BANK APO/BANK APO 2026.xlsx",
}

# ── NOMS D'ONGLETS ────────────────────────────────────────────
NOMS_MOIS_FR = {
    1: 'JANVIER', 2: 'FEVRIER', 3: 'MARS',   4: 'AVRIL',
    5: 'MAI',     6: 'JUIN',    7: 'JUILLET', 8: 'AOUT',
    9: 'SEPTEMBRE', 10: 'OCTOBRE', 11: 'NOVEMBRE', 12: 'DECEMBRE',
}
LIBELLES_FR = {
    1: 'janvier', 2: 'février', 3: 'mars',   4: 'avril',
    5: 'mai',     6: 'juin',    7: 'juillet', 8: 'août',
    9: 'septembre', 10: 'octobre', 11: 'novembre', 12: 'décembre',
}
SHEETS_PAR_MOIS = {
    "caisse_apo":  {m: "CAISSE APO"         for m in range(1, 13)},
    "caisse_apo2": {m: "CAISSE 2 APO"       for m in range(1, 13)},
    "vente_huile": {m: "VENTE D'HUILE 2026" for m in range(1, 13)},
}

def sheet_caisse_graine(mois): return f"CAISSE GRAINE {NOMS_MOIS_FR[mois]}"
def sheet_production(mois):    return f"{NOMS_MOIS_FR[mois]} 2026"
def sheet_vente_palmiste(mois):
    if mois == 1: return "VENTE NOIX PALMISTE JANVIER"
    if mois == 2: return "VEMTE NOIX DE PALMISTE FEVRIER"   # typo connue dans l'Excel
    if mois == 4: return "VENTE PALMISTE AVRIL"              # nom court en avril
    if mois == 5: return "VENTE PALMISTE MAI"                # nom court en mai
    return f"VENTE NOIX DE PALMISTE {NOMS_MOIS_FR[mois]}"
def sheet_vente_florentin(mois): return f"VENTE FLORENTIN {NOMS_MOIS_FR[mois]}"
def sheet_vente_bassin(mois):    return f"VENTE BASSIN {NOMS_MOIS_FR[mois]}"


# ── UTILITAIRES ───────────────────────────────────────────────

DRY_RUN   = False  # modifié par argparse
TENANT_ID = None   # modifié par argparse (obligatoire)

def log(msg: str):
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}", flush=True)

def categorize_libelle(libelle: str) -> str:
    """Catégorise une ligne de caisse. Miroir exact de categorizeLibelle() dans useMoisDB.js."""
    l = (libelle or "").upper()
    if any(k in l for k in ["SALAIRE", "PAIE DU", "PAIE DES", "PAIE JOUR", "PRIME POUR",
                              "PRIME DE PRODUCTION", "PRIME MOIS", "AVANCE SUR SALAIRE",
                              "JOUR FERIER", "TRAVAILLEURS TEMPORAIRE", "PESONNEL JOUR",
                              "CNPS", "CMU"]):
        return "charges_personnel"
    if any(k in l for k in ["CARBURANT", "GASOIL", "GAZOIL", "ESSENCE",
                              "ELECTRICIT", "CIE", "ACHAT EAU", "EAU POUR", "FOURNITURES",
                              "ACHAT DIVERS", "ENCRE", "FILTRE A EAU",
                              "MATERIELS MECANIQUE", "MATERIEL MECANIQUE", "SOUDURE",
                              "CHAUDIERE", "ORDINATEUR", "INFORMATIQUE", "BUREAU+ARMOIRE",
                              "MATERIEL ET EQUIPEMENT", "REFRIGERATEUR", "SPLIT", "FRIGO",
                              "STARLINK", "CIMENT", "GRAVIER", "SABLE", "FORAGE", "DALLE",
                              "BRIQUES", "CONTRE PLAQUE", "BASSIN LAGUNAGE", "TUYAUX",
                              "CONSTRUCTION", "GARAGE ENGIN", "BUREAU ANNEXE",
                              "CERTIFICAT FONCIER", "TERRAIN", "LEGALISATION"]):
        return "fournitures_usine"
    if l.startswith("ACHAT"):
        return "fournitures_usine"
    if any(k in l for k in ["VEHICULE", "BULDOZER", "BULL", "PORTE-CHAR", "PORTE CHAR",
                              "VISITE TECHNIQUE", "BILLET AVION", "LAVAGE PICK",
                              "LOCATION CAMION", "FRAIS DE TRANSPORT", "TRANSPORT FOURNISSEUR",
                              "ACHAT DE MOTO", "AFFAIRES MARITIMES", "NIVELEUSE",
                              "DEPOT DE RAFFE"]):
        return "frais_transport"
    if any(k in l for k in ["REPARATION", "REAPARATION", "ENTRETIEN", "REBOBINAGE",
                              "RECHARGEMENT BOUTEILLE", "DEPANNAGE", "DETRATAGE",
                              "NETTOYAGE", "TEFLON", "FABRICATION", "CERVEAU DE FREIN",
                              "TAMPON POUR BENNE", "DECANTEUR", "FLEXIBLE POUR CHARGEUSE",
                              "PIECES DE RECHANGE", "ASSURANCE"]):
        return "services_ext"
    if any(k in l for k in ["MAIN D'O", "MAIN DO", "MAIN DOEUVRE", "ACOMPTE SUR MAIN",
                              "SOLDE MAIN", "REBOBINEUR", "TOURNEUR", "MACONNERIE", "VITRIER",
                              "SECURIT", "GARDIEN", "VISA", "JURIDIQUE", "MEDICAUX",
                              "PRET", "FRAIS DIVERS", "MISSION", "HEBERGEMENT"]):
        return "autres_services_ext"
    if any(k in l for k in ["BAKCHICH", "BACKCHICH", "BAKHCHICH", "FRAIS RELATIONNEL",
                              "AIDE FINANCIERE", "DON POUR", "MOBILE MONEY",
                              "RELATIONNEL", "ASSISTANCE FUNEBRE", "TEE-SHORT", "POLOS"]):
        return "autres_charges"
    return "autres_services_ext"


def safe_float(val, default=0.0) -> float:
    try:   return float(val) if val is not None else default
    except (TypeError, ValueError): return default

def safe_int(val, default=0) -> int:
    try:   return int(val) if val is not None else default
    except (TypeError, ValueError): return default

def md5_bytes(content: bytes) -> str:
    return hashlib.md5(content).hexdigest()


def telecharger_dropbox(dropbox_path: str, max_retries: int = 3) -> bytes | None:
    """Télécharge un fichier Dropbox en mémoire (bytes). Retourne None si introuvable."""
    for attempt in range(1, max_retries + 1):
        try:
            _, resp = dbx.files_download(dropbox_path)
            return resp.content
        except ApiError as e:
            log(f"  ⚠️  Fichier Dropbox introuvable : {dropbox_path}")
            log(f"      Erreur : {e}")
            return None
        except Exception as e:
            if attempt < max_retries:
                wait = 15 * attempt
                log(f"  ⚠️  Erreur réseau (tentative {attempt}/{max_retries}), retry dans {wait}s : {e}")
                time.sleep(wait)
            else:
                log(f"  ❌  Échec après {max_retries} tentatives : {dropbox_path}")
                raise


def lire_sheet(content: bytes, sheet_name: str):
    """Parse un onglet Excel depuis des bytes en mémoire."""
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        log(f"  ⚠️  Sheet '{sheet_name}' introuvable (sheets dispo : {wb.sheetnames[:5]})")
        wb.close()
        return None
    lignes = list(wb[sheet_name].iter_rows(values_only=True))
    wb.close()
    return lignes


CHUNK_SIZE = 200

def inserer(table: str, rows: list, periode_id: int | None = None, label: str = ""):
    """Supprime les anciennes lignes du mois puis insère les nouvelles par blocs.
    Injecte automatiquement tenant_id dans chaque row si TENANT_ID est défini."""
    if not rows:
        log(f"  ℹ️  {table} : aucune ligne à insérer {label}")
        return
    if TENANT_ID:
        for r in rows:
            r["tenant_id"] = TENANT_ID
    if DRY_RUN:
        log(f"  [DRY] {table} : {len(rows)} lignes (période {periode_id}) {label}")
        return
    if periode_id is not None:
        q = sb.table(table).delete().eq("periode_id", periode_id)
        if TENANT_ID:
            q = q.eq("tenant_id", TENANT_ID)
        q.execute()
    for i in range(0, len(rows), CHUNK_SIZE):
        sb.table(table).insert(rows[i:i + CHUNK_SIZE]).execute()
    log(f"  ✅ {table} : {len(rows)} lignes insérées {label}")


def get_periode_id(annee: int, mois: int) -> int | None:
    q = sb.table("periodes").select("id").eq("annee", annee).eq("mois", mois)
    if TENANT_ID:
        q = q.eq("tenant_id", TENANT_ID)
    r = q.execute()
    return r.data[0]["id"] if r.data else None


def assurer_periodes() -> dict:
    """Crée les périodes manquantes dans Supabase. Retourne {mois: periode_id}.
    Exclut les mois statiques (MOIS_STATIQUES) gérés par les fichiers JS."""
    periodes = {}
    for mois in range(1, MOIS_MAX + 1):
        if mois in MOIS_STATIQUES:
            continue  # Jan/Fév/Mar → données statiques, pas Supabase
        pid = get_periode_id(ANNEE_COURANTE, mois)
        if not pid:
            log(f"  📅 Création période {LIBELLES_FR[mois]} {ANNEE_COURANTE}...")
            if not DRY_RUN:
                row = {"annee": ANNEE_COURANTE, "mois": mois, "libelle": LIBELLES_FR[mois]}
                if TENANT_ID:
                    row["tenant_id"] = TENANT_ID
                sb.table("periodes").insert(row).execute()
                pid = get_periode_id(ANNEE_COURANTE, mois)
        periodes[mois] = pid
    return periodes


# ── PARSEURS ─────────────────────────────────────────────────

def parse_production_tous_mois(content: bytes, periodes: dict):
    """Parse et insère toutes les données de production depuis le fichier de production."""
    if not DRY_RUN:
        q = sb.table("production_journaliere").delete().gte("id", 0)
        if TENANT_ID:
            q = q.eq("tenant_id", TENANT_ID)
        q.execute()
        log(f"  🗑  production_journaliere vidée (tenant: {TENANT_ID})")

    total = 0
    for mois, periode_id in periodes.items():
        lignes = lire_sheet(content, sheet_production(mois))
        if not lignes:
            continue

        rows_par_date = {}
        for row in lignes[2:]:
            if not row[1] or not isinstance(row[1], datetime):
                continue
            if row[1].month != mois:
                continue
            date_str = row[1].date().isoformat()

            if date_str not in rows_par_date:
                rows_par_date[date_str] = {
                    "periode_id":             periode_id,
                    "date_production":        date_str,
                    "regime_recu_kg":         safe_float(row[2]),
                    "regime_traite_kg":       safe_float(row[3]),
                    "huile_produite_kg":      safe_float(row[5]),
                    "livraison_citerne_kg":   safe_float(row[7]),
                    "nb_sterilisateurs":      safe_int(row[13]),
                    "livraison_florentin_kg": safe_float(row[14]),
                    "livraison_bassin_kg":    safe_float(row[15]),
                    "production_palmiste_kg": safe_float(row[16]),
                    "livraison_palmiste_kg":  safe_float(row[17]),
                    "regime_restant_kg":      safe_float(row[4]),
                    "stock_huile_kg":         safe_float(row[8]),
                    "tank_1000_kg":           safe_float(row[9]),
                    "tank_300_kg":            safe_float(row[10]),
                    "stock_graine_1_kg":      safe_float(row[11]),
                    "stock_graine_2_kg":      safe_float(row[12]),
                    "stock_palmiste_kg":      safe_float(row[18]) if len(row) > 18 else 0.0,
                    "taux_extraction":        0.0,
                }
            else:
                r = rows_par_date[date_str]
                r["regime_recu_kg"]         += safe_float(row[2])
                r["regime_traite_kg"]       += safe_float(row[3])
                r["huile_produite_kg"]      += safe_float(row[5])
                r["livraison_citerne_kg"]   += safe_float(row[7])
                r["nb_sterilisateurs"]      += safe_int(row[13])
                r["livraison_florentin_kg"] += safe_float(row[14])
                r["livraison_bassin_kg"]    += safe_float(row[15])
                r["production_palmiste_kg"] += safe_float(row[16])
                r["livraison_palmiste_kg"]  += safe_float(row[17])
                r["regime_restant_kg"]       = safe_float(row[4])
                r["stock_huile_kg"]          = safe_float(row[8])
                r["tank_1000_kg"]            = safe_float(row[9])
                r["tank_300_kg"]             = safe_float(row[10])
                r["stock_graine_1_kg"]       = safe_float(row[11])
                r["stock_graine_2_kg"]       = safe_float(row[12])
                if len(row) > 18:
                    r["stock_palmiste_kg"]   = safe_float(row[18])

        for r in rows_par_date.values():
            traite = r["regime_traite_kg"]
            r["taux_extraction"] = round(r["huile_produite_kg"] / traite, 4) if traite else 0.0

        rows = list(rows_par_date.values())
        if rows and not DRY_RUN:
            if TENANT_ID:
                for r in rows:
                    r["tenant_id"] = TENANT_ID
            for i in range(0, len(rows), CHUNK_SIZE):
                sb.table("production_journaliere").insert(rows[i:i + CHUNK_SIZE]).execute()
            log(f"  ✅ production_journaliere : {len(rows)} lignes (mois {mois})")
            total += len(rows)
        elif rows:
            log(f"  [DRY] production_journaliere : {len(rows)} lignes (mois {mois})")

    return total


def parse_ventes_huile(content: bytes, mois: int, periode_id: int):
    lignes = lire_sheet(content, SHEETS_PAR_MOIS["vente_huile"][mois])
    if not lignes: return

    q = sb.table("clients").select("id").eq("reference", "SARCI")
    if TENANT_ID:
        q = q.eq("tenant_id", TENANT_ID)
    sarci = q.single().execute()
    client_id = sarci.data["id"] if sarci.data else None

    rows = []
    for row in lignes:
        if not row[1] or not isinstance(row[1], datetime): continue
        if row[1].month != mois: continue
        # Tolérer erreurs de saisie année ±1 (ex: 2025-02-26 au lieu de 2026-02-26)
        if abs(row[1].year - ANNEE_COURANTE) > 1: continue
        if safe_float(row[3]) == 0: continue
        # Forcer l'année correcte pour date_vente en cas d'erreur de saisie
        date_corrigee = row[1].replace(year=ANNEE_COURANTE)
        rows.append({
            "periode_id":     periode_id,
            "client_id":      client_id,
            "date_vente":     date_corrigee.date().isoformat(),
            "libelle":        str(row[2] or ""),
            "poids_apo_kg":   safe_float(row[3]),
            "poids_sarci_kg": safe_float(row[4]),
            "prix_kg":        safe_float(row[6]),
            "avance_sarci":   safe_float(row[7]),
        })
    inserer("ventes_huile", rows, periode_id, f"(mois {mois})")


def parse_ventes_palmiste(content: bytes, mois: int, periode_id: int):
    lignes = lire_sheet(content, sheet_vente_palmiste(mois))
    if not lignes: return

    rows = []
    for row in lignes[1:]:
        if not row[1] or not isinstance(row[1], datetime): continue
        if safe_float(row[3]) == 0: continue
        rows.append({
            "periode_id": periode_id,
            "date_vente": row[1].date().isoformat(),
            "poids_kg":   safe_float(row[3]),
            "prix_kg":    safe_float(row[4]) or 60.0,
        })
    inserer("ventes_palmiste", rows, periode_id, f"(mois {mois})")


def parse_ventes_florentin(content: bytes, mois: int, periode_id: int):
    lignes = lire_sheet(content, sheet_vente_florentin(mois))
    if not lignes: return

    rows = []
    for row in lignes[1:]:
        if not row[1] or not isinstance(row[1], datetime): continue
        if safe_float(row[3]) == 0: continue
        rows.append({
            "periode_id": periode_id,
            "date_vente": row[1].date().isoformat(),
            "poids_kg":   safe_float(row[3]),
            "prix_kg":    safe_float(row[4]),
        })
    inserer("ventes_florentin", rows, periode_id, f"(mois {mois})")


def parse_ventes_bassin(content: bytes, mois: int, periode_id: int):
    lignes = lire_sheet(content, sheet_vente_bassin(mois))
    if not lignes: return

    rows = []
    for row in lignes[1:]:
        if not row[1] or not isinstance(row[1], datetime): continue
        if safe_float(row[3]) == 0: continue
        rows.append({
            "periode_id":   periode_id,
            "date_vente":   row[1].date().isoformat(),
            "poids_kg":     safe_float(row[3]),
            "prix_kg":      safe_float(row[4]) or None,
            "facture":      safe_float(row[4]) > 0,
            "montant_fcfa": safe_float(row[5]),
        })
    inserer("ventes_bassin", rows, periode_id, f"(mois {mois})")


def parse_caisse_apo(content: bytes, mois: int, periode_id: int):
    lignes = lire_sheet(content, SHEETS_PAR_MOIS["caisse_apo"][mois])
    if not lignes: return

    SKIP = ("TRANSFERT", "VIREMENT", "VERSEMENT", "DEPOT", "APPRO")
    rows = []
    for row in lignes:
        if not row[1] or not isinstance(row[1], datetime): continue
        if row[1].month != mois: continue
        libelle = str(row[2] or "")
        rows.append({
            "periode_id":     periode_id,
            "date_mouvement": row[1].date().isoformat(),
            "libelle":        libelle,
            "debit_fcfa":     safe_float(row[3]),
            "credit_fcfa":    safe_float(row[4]),
            "solde_fcfa":     safe_float(row[5]),
            "categorie":      categorize_libelle(libelle),
        })
    inserer("caisse_apo", rows, periode_id, f"(mois {mois})")


def parse_caisse_apo2(content: bytes, mois: int, periode_id: int):
    lignes = lire_sheet(content, SHEETS_PAR_MOIS["caisse_apo2"][mois])
    if not lignes: return

    rows = []
    for row in lignes:
        if not row[1] or not isinstance(row[1], datetime): continue
        if row[1].month != mois: continue
        libelle = str(row[2] or "")
        rows.append({
            "periode_id":     periode_id,
            "date_mouvement": row[1].date().isoformat(),
            "libelle":        libelle,
            "debit_fcfa":     safe_float(row[3]),
            "credit_fcfa":    safe_float(row[4]),
            "solde_fcfa":     safe_float(row[5]),
            "categorie":      categorize_libelle(libelle),
        })
    inserer("caisse_apo2", rows, periode_id, f"(mois {mois})")


def parse_achats_regimes(content: bytes, mois: int, periode_id: int):
    lignes = lire_sheet(content, sheet_caisse_graine(mois))
    if not lignes: return

    refs_vues = set()
    for row in lignes[1:]:
        if row[3] and isinstance(row[3], datetime):
            refs_vues.add(str(row[4] or "INCONNU").strip())

    if not DRY_RUN:
        for ref in refs_vues:
            row = {"reference": ref, "nom": ref}
            if TENANT_ID:
                row["tenant_id"] = TENANT_ID
            sb.table("fournisseurs").upsert(row, on_conflict="reference,tenant_id").execute()

    fournisseurs_map = {}
    for ref in refs_vues:
        q = sb.table("fournisseurs").select("id").eq("reference", ref)
        if TENANT_ID:
            q = q.eq("tenant_id", TENANT_ID)
        r = q.single().execute()
        if r.data:
            fournisseurs_map[ref] = r.data["id"]

    rows = []
    for row in lignes[1:]:
        if not row[3] or not isinstance(row[3], datetime): continue
        if safe_float(row[5]) == 0: continue
        ref = str(row[4] or "INCONNU").strip()
        rows.append({
            "periode_id":     periode_id,
            "fournisseur_id": fournisseurs_map.get(ref),
            "date_achat":     row[3].date().isoformat(),
            "type_transport": str(row[1] or ""),
            "numero_camion":  str(row[2] or ""),
            "poids_kg":       safe_float(row[5]),
            "prix_kg":        safe_float(row[6]),
            "prix_transport": safe_float(row[7]),
            "appro_caisse":   safe_float(row[8]),
            # montant_total est une colonne générée (prix_kg × poids_kg) — ne pas l'insérer
        })
    inserer("achats_regimes", rows, periode_id, f"(mois {mois})")


def parse_pepiniere(content: bytes):
    lignes = lire_sheet(content, "Feuil1")
    if not lignes: return

    rows_clients, rows_contrats = [], []
    for row in lignes[1:]:
        ordre = str(row[0] or "").strip()
        if not ordre.isdigit(): continue
        nom = str(row[2] or "").strip()
        if not nom: continue
        rows_clients.append({"reference": nom[:50], "nom": nom, "telephone": str(row[3] or ""), "type": "pepiniere"})
        date_contrat = row[1].date().isoformat() if isinstance(row[1], datetime) else None
        if not date_contrat: continue
        rows_contrats.append({
            "numero_ordre":        ordre,
            "date_contrat":        date_contrat,
            "localite_champ":      str(row[4] or ""),
            "superficie_demandee": safe_float(row[5]),
            "superficie_champ_ha": safe_float(row[7]),
            "prix_unitaire_fcfa":  safe_float(row[6]),
            "montant_total":       safe_float(row[8]),
            "net_encaisse":        safe_float(row[9]),
            "_nom_client":         nom[:50],
        })

    if not DRY_RUN:
        for rc in rows_clients:
            if TENANT_ID:
                rc["tenant_id"] = TENANT_ID
            sb.table("clients").upsert(rc, on_conflict="reference,tenant_id").execute()

        q = sb.table("contrats_pepiniere").delete().neq("id", 0)
        if TENANT_ID:
            q = q.eq("tenant_id", TENANT_ID)
        q.execute()

        for rc in rows_contrats:
            nom_ref = rc.pop("_nom_client")
            cq = sb.table("clients").select("id").eq("reference", nom_ref)
            if TENANT_ID:
                cq = cq.eq("tenant_id", TENANT_ID)
            cl = cq.single().execute()
            rc["client_id"] = cl.data["id"] if cl.data else None
            if TENANT_ID:
                rc["tenant_id"] = TENANT_ID
            sb.table("contrats_pepiniere").insert(rc).execute()

    log(f"  ✅ contrats_pepiniere : {len(rows_contrats)} lignes insérées")


_SKIP_BANQUE = ("RETRAIT ESPECES", "APPROV CAISSE", "DEPOT ESPECES POUR APPROV")

# Catégories banque (CHECK constraint Supabase) :
# amortissement, assurance, charges_patronales, construction,
# electricite, entretien, frais_admin, frais_bancaires,
# frais_relat, materiels, salaires, securite, taxes_fiscales, vehicules

def categorize_banque(libelle: str) -> str:
    l = (libelle or "").upper()
    if any(k in l for k in ["PRÊT BANCAIRE", "PRET BANCAIRE", "ECHEANCE PRÊ", "ECHEANCE PRET",
                              "AMORTISSEMENT PRET", "ECH PRET"]):
        return "amortissement"
    if any(k in l for k in ["ASSURANCE", "SUNU ASSURANCE", "WAFA ASSURANCE"]):
        return "assurance"
    if any(k in l for k in ["CNPS", "CNAM", "IPS-CNPS", "IPS-CNAM", "IPS CNPS",
                              "COTISATION CNPS", "COTISATION CNAM"]):
        return "charges_patronales"
    if any(k in l for k in ["CONSTRUCTION", "FORAGE", "DALLE", "CIMENT", "GRAVIER", "BRIQUES",
                              "BASSIN LAGUNAGE", "TRAVAUX AMENAGEMENT", "STATION D'EPURATION",
                              "ERT-B", "EPURATION"]):
        return "construction"
    if any(k in l for k in ["CIE/", "/CIE/", "CIE ", "CONSOMMATION ELECTRICIT", "ELECTRICIT"]):
        return "electricite"
    if any(k in l for k in ["ENTRETIEN", "REPARATION", "REAPARATION", "REBOBINAGE", "SOUDURE",
                              "MAINTENANCE", "DEPANNAGE", "DETRATAGE", "NETTOYAGE", "CHAUDIERE"]):
        return "entretien"
    if any(k in l for k in ["CABINET BLANCHARD", "ASSISTANCE COMPTABLE", "NP CONSEILS",
                              "NOTAIRE", "JURIDIQUE", "CONSULTANT", "EXPERT COMPTABLE"]):
        return "frais_admin"
    if any(k in l for k in ["AGIOS", "FRAIS BDA", "FRAIS BANCAIRE", "TENUE DE COMPTE",
                              "RETOUR,INT", "TAXE FRAIS FIXE", "FRAIS FINANCIER",
                              "COMMISSION", "IBE PREMIUM", "PACK IBE"]):
        return "frais_bancaires"
    if any(k in l for k in ["CHPH", "AIPH", "COTISATION FONCTIONNEMENT", "REDEVANCE CONSEIL",
                              "FAMILLE PROFESSIONNELLE", "DEVELOPPEMENT FILIERE",
                              "INTERPROFESSION"]):
        return "frais_relat"
    if any(k in l for k in ["MATERIEL", "FOURNITURE", "DMD SARL", "PIECES", "POMPE", "VANNE",
                              "TUYAUX", "FLEXIB", "TM&FRERES", "ENIGMA", "MONDE CYCLES",
                              "MONDIAL CYCLES", "SEMAG MATFORCE", "PALMITECO",
                              "SN SATA", "SN ROCK"]):
        return "materiels"
    if any(k in l for k in ["VIREMENT SALAIRE", "SALAIRE ", "PAIE DU", "PAIE DES"]):
        return "salaires"
    if any(k in l for k in ["GARDIENNAGE", "POWERMAX", "SECURIT", "GARDE"]):
        return "securite"
    if any(k in l for k in ["TELEPAIEMENT ITS", "TELEPAIEMENT TSE", "ITS-FDFP", "RIBIC", "FIRCA",
                              "TSE ", "REGIE DES MINES", "TAXE CHPH", "CMU", "TAXE D'INSPECTION",
                              "TAXE ANNUELLE", "IMPOT", "TAXE "]):
        return "taxes_fiscales"
    if any(k in l for k in ["VEHICULE", "CAMION", "TRANSPORT", "UNICARS", "MOTO", "BULDOZER",
                              "PORTE-CHAR", "PORTE CHAR", "AFFAIRES MARITIMES", "BILLET",
                              "LIVRAISON", "MEITE LADJI"]):
        return "vehicules"
    return "materiels"  # fallback dans la liste autorisée


def parse_banque_apo(content: bytes, mois: int, periode_id: int):
    """Parse SGCI GL APO et BDA GL APO depuis BANK APO 2026.xlsx.

    SGCI: col0=date_op, col3=libellé, col4=débit, col7=date_valeur (pour filtre mois).
    BDA format A: col3=libellé (str), col4=débit, col6=date_valeur.
    BDA format B: col0=date_op, col1=libellé, col3=montant négatif, col6=date_valeur.
    """
    rows_out = []

    # ── SGCI GL APO ───────────────────────────────────────────────
    sgci = lire_sheet(content, "SGCI GL APO")
    if sgci:
        for row in sgci:
            date_val = row[7] if len(row) > 7 and isinstance(row[7], datetime) else None
            date_op  = row[0] if isinstance(row[0], datetime) else None
            date_ref = date_val or date_op
            if not date_ref or date_ref.year != ANNEE_COURANTE or date_ref.month != mois:
                continue
            debit = safe_float(row[4] if len(row) > 4 else None)
            if debit <= 0:
                continue
            libelle = str(row[3] or "").strip()
            if not libelle:
                continue
            if any(k in libelle.upper() for k in _SKIP_BANQUE):
                continue
            rows_out.append({
                "periode_id":     periode_id,
                "banque":         "SGCI",
                "date_operation": (date_op or date_val).date().isoformat(),
                "date_valeur":    (date_val or date_op).date().isoformat(),
                "libelle":        libelle,
                "montant_fcfa":   int(round(debit)),
                "categorie":      categorize_banque(libelle),
            })

    # ── BDA GL APO ────────────────────────────────────────────────
    bda = lire_sheet(content, "BDA GL APO")
    if bda:
        for row in bda:
            date_val = row[6] if len(row) > 6 and isinstance(row[6], datetime) else None
            date_op  = row[0] if isinstance(row[0], datetime) else None
            date_ref = date_val or date_op
            if not date_ref or date_ref.year != ANNEE_COURANTE or date_ref.month != mois:
                continue

            # Format A : libellé en col3 (str), débit en col4
            if isinstance(row[3], str) and row[3].strip():
                libelle = row[3].strip()
                debit   = safe_float(row[4] if len(row) > 4 else None)
            # Format B : libellé en col1, montant négatif en col3
            elif isinstance(row[3], (int, float)) and row[3] < 0 and isinstance(row[1], str) and row[1].strip():
                libelle = row[1].strip()
                debit   = abs(safe_float(row[3]))
            else:
                continue

            if debit <= 0:
                continue
            lu = libelle.upper()
            if lu.startswith(("TOTAL", "SOLDE", "JANVIER", "FEVRIER", "MARS",
                               "AVRIL", "MAI", "JUIN", "JUILLET", "AOUT",
                               "SEPTEMBRE", "OCTOBRE", "NOVEMBRE", "DECEMBRE")):
                continue
            rows_out.append({
                "periode_id":     periode_id,
                "banque":         "BDA",
                "date_operation": (date_op or date_val).date().isoformat(),
                "date_valeur":    (date_val or date_op).date().isoformat(),
                "libelle":        libelle,
                "montant_fcfa":   int(round(debit)),
                "categorie":      categorize_banque(libelle),
            })

    inserer("banque_apo", rows_out, periode_id, f"(mois {mois})")


def recalculer_kpis(mois: int, periode_id: int):
    """Recalcule kpis_mensuels depuis les tables de détail."""
    vh  = sb.table("ventes_huile").select("montant_fcfa").eq("periode_id", periode_id).execute()
    vp  = sb.table("ventes_palmiste").select("montant_fcfa").eq("periode_id", periode_id).execute()
    vf  = sb.table("ventes_florentin").select("montant_fcfa").eq("periode_id", periode_id).execute()
    vb  = sb.table("ventes_bassin").select("montant_fcfa").eq("periode_id", periode_id).execute()
    ca1 = sb.table("caisse_apo").select("credit_fcfa,libelle").eq("periode_id", periode_id).execute()
    ca2 = sb.table("caisse_apo2").select("credit_fcfa,libelle").eq("periode_id", periode_id).execute()
    bq  = sb.table("banque_apo").select("montant_fcfa,categorie").eq("periode_id", periode_id).execute()
    pj  = sb.table("production_journaliere").select("*").eq("periode_id", periode_id).execute()
    ar  = sb.table("achats_regimes").select("poids_kg,prix_kg,prix_transport").eq("periode_id", periode_id).execute()
    pep_q = sb.table("contrats_pepiniere").select("montant_total,net_encaisse")
    if TENANT_ID:
        pep_q = pep_q.eq("tenant_id", TENANT_ID)
    pep = pep_q.execute()

    # Patterns de transferts inter-caisses à exclure des charges
    SKIP = ("TRANSFERT", "VIREMENT", "VERSEMENT", "DEPOT", "APPRO")

    def is_transfer(libelle: str) -> bool:
        u = (libelle or "").upper()
        return any(u.startswith(k) for k in SKIP)

    ca_huile     = sum(safe_float(r["montant_fcfa"]) for r in vh.data)
    ca_palmiste  = sum(safe_float(r["montant_fcfa"]) for r in vp.data)
    ca_florentin = sum(safe_float(r["montant_fcfa"]) for r in vf.data)
    ca_bassin    = sum(safe_float(r["montant_fcfa"]) for r in vb.data)

    # Charges caisse (hors transferts inter-caisses)
    charges_caisse = (
        sum(safe_float(r["credit_fcfa"]) for r in ca1.data if not is_transfer(r["libelle"]))
      + sum(safe_float(r["credit_fcfa"]) for r in ca2.data if not is_transfer(r["libelle"]))
    )

    # Charges banque — séparées par type
    CATS_FIN = {"amortissement", "frais_bancaires"}
    charges_banque = sum(safe_float(r["montant_fcfa"]) for r in (bq.data or [])
                         if r.get("categorie") not in CATS_FIN)
    amort_banque   = sum(safe_float(r["montant_fcfa"]) for r in (bq.data or [])
                         if r.get("categorie") == "amortissement")
    frais_fin      = sum(safe_float(r["montant_fcfa"]) for r in (bq.data or [])
                         if r.get("categorie") == "frais_bancaires")

    # Totaux
    charges      = charges_caisse + charges_banque
    amort_total  = amort_banque or 0   # banque = source prioritaire
    frais_total  = frais_fin

    reg_recus   = sum(safe_float(r["regime_recu_kg"])        for r in pj.data)
    reg_traites = sum(safe_float(r["regime_traite_kg"])       for r in pj.data)
    huile_prod  = sum(safe_float(r["huile_produite_kg"])      for r in pj.data)
    huile_vend  = sum(safe_float(r["livraison_citerne_kg"])   for r in pj.data)
    palm_prod   = sum(safe_float(r["production_palmiste_kg"]) for r in pj.data)
    palm_vend   = sum(safe_float(r["livraison_palmiste_kg"])  for r in pj.data)
    nb_steril   = sum(safe_int(r["nb_sterilisateurs"])        for r in pj.data)
    stock_fin   = safe_float(pj.data[-1]["regime_restant_kg"]) if pj.data else 0

    total_poids = sum(safe_float(r["poids_kg"]) for r in ar.data)
    # coût MP = (prix_kg + prix_transport) × poids_kg — transport inclus comme dans l'Excel col9
    total_mont  = sum(
        safe_float(r["poids_kg"]) * (safe_float(r["prix_kg"]) + safe_float(r["prix_transport"]))
        for r in ar.data
    )
    prix_moy    = total_mont / total_poids if total_poids else 0
    nb_camions  = len(ar.data)

    prix_huile  = ca_huile / huile_vend if huile_vend else 0
    te          = huile_prod / reg_traites if reg_traites else 0
    # Coût MP = graines nécessaires pour produire l'huile effectivement vendue
    # graine_vendu = huile_vendue / TE  →  cout_mp = graine_vendu × prix_moy
    graine_vendu = huile_vend / te if te else 0
    cout_mp      = graine_vendu * prix_moy
    ca_total    = ca_huile + ca_palmiste + ca_florentin + ca_bassin
    resultat    = ca_total - cout_mp - charges - amort_total - frais_total
    marge       = round(resultat / ca_total * 100, 2) if ca_total else 0

    pep_total    = sum(safe_float(r["montant_total"]) for r in pep.data)
    pep_encaisse = sum(safe_float(r["net_encaisse"])  for r in pep.data)

    payload = {
        "periode_id":              periode_id,
        **({"tenant_id": TENANT_ID} if TENANT_ID else {}),
        "ca_huile_fcfa":           ca_huile,
        "ca_palmiste_fcfa":        ca_palmiste,
        "ca_florentin_fcfa":       ca_florentin,
        "ca_bassin_fcfa":          ca_bassin,
        "cout_mp_fcfa":            cout_mp,
        "charges_exploitation":    charges,
        "amortissement_fcfa":      amort_total + frais_total,
        "marge_nette_pct":         marge,
        "regimes_recus_kg":        reg_recus,
        "regimes_traites_kg":      reg_traites,
        "stock_fin_mois_kg":       stock_fin,
        "huile_produite_kg":       huile_prod,
        "huile_vendue_kg":         huile_vend,
        "taux_extraction":         round(te, 4),
        "nb_camions":              nb_camions,
        "nb_sterilisateurs":       nb_steril,
        "palmiste_produit_kg":     palm_prod,
        "palmiste_vendu_kg":       palm_vend,
        "prix_moyen_regime_kg":    round(prix_moy, 2),
        "prix_moyen_huile_kg":     round(prix_huile, 2),
        "pepiniere_contrats_fcfa": pep_total,
        "pepiniere_encaisse_fcfa": pep_encaisse,
        "pepiniere_restant_fcfa":  pep_total - pep_encaisse,
        "mis_a_jour_le":           datetime.now().isoformat(),
    }

    if not DRY_RUN:
        sb.table("kpis_mensuels").upsert(payload, on_conflict="periode_id").execute()
    log(f"  ✅ KPIs mois {mois} — CA : {ca_total:,.0f} | Charges : {charges:,.0f} | Amort : {amort_total:,.0f} | Résultat : {resultat:,.0f} FCFA")


# ── ORCHESTRATEUR ─────────────────────────────────────────────

def run(mois_cible: int | None = None):
    """
    mois_cible : si fourni, n'importe que ce mois précis.
                 Sinon importe tous les mois de 1 à MOIS_MAX.
    """
    log("=" * 56)
    mois_min_dyn = min((m for m in range(1, 13) if m not in MOIS_STATIQUES), default=1)
    statiques_info = f"mois 1-{max(MOIS_STATIQUES)} statiques" if MOIS_STATIQUES else "tous mois Supabase"
    log(f"APO ETL Cloud — {LIBELLES_FR[mois_min_dyn]} → {LIBELLES_FR[MOIS_MAX]} {ANNEE_COURANTE} ({statiques_info})")
    if DRY_RUN:
        log("⚠️  MODE DRY-RUN — aucune écriture Supabase")
    log("=" * 56)

    # 1. Périodes Supabase
    periodes = assurer_periodes()
    if mois_cible:
        periodes = {mois_cible: periodes[mois_cible]}
    log(f"  Mois traités : {list(periodes.keys())}")

    # 2. Téléchargement Dropbox — 1 fichier = 1 download
    log("─" * 56)
    log("📥 Téléchargement des fichiers Dropbox...")
    fichiers = {}
    for cle, dropbox_path in FICHIERS_DROPBOX.items():
        log(f"  ↓ {cle} ...")
        content = telecharger_dropbox(dropbox_path)
        if content:
            fichiers[cle] = content
            log(f"    {len(content) / 1024:.0f} KB téléchargés")
        else:
            log(f"  ❌ {cle} ignoré (fichier manquant)")

    # 3. Import Production (tous les mois d'un coup)
    if "production" in fichiers:
        log("─" * 56)
        log("🏭 Import production...")
        try:
            parse_production_tous_mois(fichiers["production"], periodes)
        except Exception as e:
            log(f"  ❌ Erreur production : {e}")

    # 4. Pépinière (table cumulative, pas de boucle mois)
    if "pepiniere" in fichiers:
        log("─" * 56)
        log("🌱 Import pépinière...")
        try:
            parse_pepiniere(fichiers["pepiniere"])
        except Exception as e:
            log(f"  ❌ Erreur pépinière : {e}")

    # 5. Sources mensuelles
    sources_mensuelles = [
        ("vente_huile",     parse_ventes_huile),
        ("vente_palmiste",  parse_ventes_palmiste),
        ("vente_florentin", parse_ventes_florentin),
        ("vente_bassin",    parse_ventes_bassin),
        ("caisse_apo",      parse_caisse_apo),
        ("caisse_apo2",     parse_caisse_apo2),
        ("caisse_graine",   parse_achats_regimes),
        ("banque_apo",      parse_banque_apo),
    ]

    for cle, parser in sources_mensuelles:
        if cle not in fichiers:
            continue
        log("─" * 56)
        log(f"📊 Import {cle}...")
        for mois, periode_id in periodes.items():
            try:
                parser(fichiers[cle], mois, periode_id)
            except Exception as e:
                log(f"  ❌ Erreur mois {mois} [{cle}] : {e}")

    # 6. Recalcul KPIs
    log("─" * 56)
    log("📈 Recalcul des KPIs mensuels...")
    for mois, pid in periodes.items():
        try:
            recalculer_kpis(mois, pid)
        except Exception as e:
            log(f"  ❌ Erreur KPIs mois {mois} : {e}")

    log("=" * 56)
    log("ETL Cloud terminé ✅")
    log("=" * 56)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="APO ETL Cloud — Dropbox → Supabase")
    parser.add_argument("--tenant", required=True, help="Identifiant du tenant (ex: apo, huilerie_xyz)")
    parser.add_argument("--mois",   type=int,      help="Importer seulement ce mois (1-12)")
    parser.add_argument("--dry",    action="store_true", help="Simulation — ne pas écrire dans Supabase")
    args = parser.parse_args()

    DRY_RUN   = args.dry
    TENANT_ID = args.tenant

    if args.mois and args.mois in MOIS_STATIQUES:
        print(f"❌ Mois {args.mois} ({LIBELLES_FR[args.mois]}) est géré par les fichiers statiques JS — ETL non nécessaire.")
        exit(1)
    run(mois_cible=args.mois)
