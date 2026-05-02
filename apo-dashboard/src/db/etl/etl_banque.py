"""
ETL — BANQUE APO 2026 (SGCI + BDA)
====================================
• Crée la table banque_apo si elle n'existe pas
• Parse BANQUE APO 26.xlsx (sheets SGCI GL APO, BDA GL APO)
• Catégorise chaque débit et l'insère en base
• Exclut : virements internes, crédits SARCI, reversements

Catégories :
  Existantes caisse : salaires, carburant, main_oeuvre, entretien, construction,
                      vehicules, materiels, eau_fournitures, frais_relat, frais_admin, autre
  Nouvelles banque  : electricite, assurance, securite, charges_patronales,
                      taxes_fiscales, frais_bancaires
  Spéciale          : amortissement  (→ Section IV P&L, pas charges exploit.)

Usage : python3 etl_banque.py [chemin_excel]
"""

import os, sys
from datetime import datetime, date
import openpyxl
import psycopg2
from dotenv import load_dotenv

load_dotenv(os.path.join(os.path.dirname(__file__), '.env'))

EXCEL_PATH  = sys.argv[1] if len(sys.argv) > 1 else os.path.expanduser('~/Desktop/BANQUE APO 26.xlsx')
DB_HOST     = os.getenv('DB_HOST', 'db.iwfgvhenqzdutjcxhuip.supabase.co')
DB_PASSWORD = os.getenv('DB_PASSWORD')
DB_USER     = 'postgres'
DB_NAME     = 'postgres'
DB_PORT     = 5432

ALL_CATEGORIES = (
    'salaires', 'carburant', 'main_oeuvre', 'entretien', 'construction',
    'vehicules', 'materiels', 'eau_fournitures', 'frais_relat', 'frais_admin',
    'electricite', 'assurance', 'securite', 'charges_patronales',
    'taxes_fiscales', 'frais_bancaires', 'amortissement', 'autre',
)

# ── CATÉGORISEUR ─────────────────────────────────────────────────────────────

def categorize_banque(libelle: str) -> str:
    """
    Retourne la catégorie d'une ligne de banque ou '_skip' si elle doit être ignorée.
    Aligné avec CAT_LABELS dans useMoisDB.js.
    """
    l = (libelle or '').upper().strip()

    # ── À ignorer ─────────────────────────────────────────────────────────────
    # Appros internes (la dépense est déjà enregistrée dans caisse_apo)
    if 'APPRO CAISSE' in l:
        return '_skip'
    # Appros SARCI (avance sur livraisons futures, pas une charge)
    if 'APPRO SARCI' in l:
        return '_skip'
    # Reversement d'un chèque impayé (annule un crédit)
    if 'COMPENSATION CHQ IMPAYEE' in l:
        return '_skip'
    # Lignes techniques des relevés (totaux, soldes)
    skip_markers = ['SOLDE FINAL', 'SOLDE INITIAL', 'TOTAL MOUVEMENTS',
                    'SOLDE INTERMEDIAIRE', 'TOTAL DEPENSES', 'ECART']
    if any(m in l for m in skip_markers):
        return '_skip'

    # ── Amortissement (Section IV) ────────────────────────────────────────────
    if 'ECHEANCE' in l and any(k in l for k in ['PRÊT', 'PRET', 'BANCAIRE']):
        return 'amortissement'

    # ── Salaires directs (virement bancaire brut) ─────────────────────────────
    if 'VIREMENT' in l and 'SALAIRE' in l:
        return 'salaires'

    # ── Charges patronales (CNPS, CMU / CNAM) ────────────────────────────────
    if any(k in l for k in ['CNPS', 'IPS-CNAM', '/CNAM/', 'CMU ']):
        return 'charges_patronales'

    # ── Électricité (CIE) ─────────────────────────────────────────────────────
    if '/CIE/' in l or '/ CIE /' in l or (' CIE' in l and ('CONSOMMATION' in l or 'FACT' in l)):
        return 'electricite'

    # ── Taxes & impôts ────────────────────────────────────────────────────────
    taxes_keys = [
        'ITS-FDFP', 'ITS FDFP', 'RIBIC', 'FIRCA', 'TSE',
        'TELEPAIEMENT ITS', 'TELEPAIEMENT TSE', 'TELEPAIEMENT IMPOT',
        'IMPOT FONCIER', 'IMPOTS BIC', 'BIC 1ER TIERS', 'PATENTE',
        'PENALITES', "REDEVANCE D'OCCUPATION", 'REDEVANCE D\u2019OCCUPATION',
        'TIERS 2025',
    ]
    if any(k in l for k in taxes_keys):
        return 'taxes_fiscales'

    # ── Sécurité / gardiennage ────────────────────────────────────────────────
    if any(k in l for k in ['POWERMAX', 'POWER MAX', 'GARDIENNAGE', 'PMS GARDIENNAGE']):
        return 'securite'

    # ── Assurances ────────────────────────────────────────────────────────────
    if any(k in l for k in ['ASSURANCE', 'WAFA', 'GNA CI']):
        return 'assurance'

    # ── Frais bancaires ───────────────────────────────────────────────────────
    bancaires_keys = [
        'AGIOS', 'PACK IBE', 'FRAIS BDA', 'TENUE DE COMPTE', 'TAXE FRAIS FIXE',
        'FACTURATION PACK', 'DEMANDE DE FDI', 'FRAIS FDI', 'FRAIS FINANCIER',
        'Retour,Int',
    ]
    if any(k in l for k in bancaires_keys):
        return 'frais_bancaires'

    # ── Frais administratifs ──────────────────────────────────────────────────
    admin_keys = [
        'CABINET BLANCHARD', 'NP CONSEILS', 'CABINET ALICA', 'ANAE',
        'AGENCE NATIONALE DE L', 'BILAN FISCAL', 'AUDIT ENVIRONNEMENTAL',
        'ASSISTANCE COMPTABLE', "FIN D'EXERCICE", 'DECLARATIONS ITS',
        'SERVICES ET IMPRIMERIES',
    ]
    if any(k in l for k in admin_keys):
        return 'frais_admin'

    # ── Cotisations interprofessionnelles ─────────────────────────────────────
    if any(k in l for k in ['CHPH', 'FONCTIONNEMENT AIPH', 'FAMILLE PROFESSIONNELLE',
                              'DEVELOPPEMENT FILIERE', 'REDEVANCE CONSEIL HEVEA']):
        return 'frais_relat'

    # ── Véhicules ────────────────────────────────────────────────────────────
    if any(k in l for k in ['UNICARS', 'VEHICULE', 'SOCIETE RAHAL', 'MEITE LADJI',
                              'TRANSPORT PIECES']):
        return 'vehicules'

    # ── Entretien & réparation ────────────────────────────────────────────────
    entretien_keys = [
        'TM&FRERES', 'ETS TM', 'SAM&SONS', 'O RING', 'GARNITURE',
        'SPRIINT TECH', 'JOINT COTIERE', 'POMPE ASPIRATEUR',
        'ENTRETIEN DES ORDINATEURS', 'ENTRETIEN ORDINATEURS',
        'ELECTROTECH',   # electrical installation → construction/entretien
    ]
    if any(k in l for k in entretien_keys):
        # Travaux d'électricité solaire/pont bascule → construction
        if 'ELECTRICITE SOLAIRE' in l or 'PONT BASCULE' in l:
            return 'construction'
        return 'entretien'

    # ── Matériels & équipements ───────────────────────────────────────────────
    materiels_keys = [
        'MONDIAL CYCLES', 'DMD SARL', 'ENIGMA TECHNOLOGIE',
        'BERNABE CI', 'BALANCE DE PRECISION', 'DISJONCTEUR',
        "MATERIEL ET FOURNITURE D'USINE", "MATERIEL ET FOUNITURE D'USINE",
        'ETAGERE', 'CLE A CHOC', 'POSTE A SOUDER', 'BARRE DE TOLE',
        'BARRE CORNIERES', 'BARRE DE FER',
    ]
    if any(k in l for k in materiels_keys):
        return 'materiels'

    # COGES : tuyaux → construction, sinon matériels
    if 'COGES' in l:
        return 'construction' if 'TUYAUX' in l else 'materiels'

    # ── Eau & fournitures ─────────────────────────────────────────────────────
    if any(k in l for k in ['IVOIRE DIGITAL', 'EXPERTISE CHIMIQUE',
                              'ENCRE', 'IMPRIMANTE', 'HYDROXYDE', 'ALCOOL ETHILIQUE',
                              'BECHER', 'CHRISTALLISOIR']):
        return 'eau_fournitures'

    # ── Construction ─────────────────────────────────────────────────────────
    if any(k in l for k in ['TUYAUX PVC', 'PONT BASCULE']):
        return 'construction'

    return 'frais_admin'   # défaut banque → frais administratifs

# ── PARSING EXCEL ─────────────────────────────────────────────────────────────

MOIS_LABELS = {
    'JANVIER': 1, 'FEVRIER': 2, 'FÉVRIER': 2, 'MARS': 3,
    'AVRIL': 4,  'MAI': 5,     'JUIN': 6,
    'JUILLET': 7,'AOUT': 8,    'AOÛT': 8, 'SEPTEMBRE': 9,
    'OCTOBRE': 10,'NOVEMBRE': 11,'DECEMBRE': 12,'DÉCEMBRE': 12,
}

def detect_month_from_header(text: str):
    """Retourne le numéro de mois si la cellule est un en-tête de section."""
    t = (text or '').upper().strip().rstrip('.')
    for label, num in MOIS_LABELS.items():
        if t.startswith(label):
            return num
    return None

def safe_date(val):
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    return None

def parse_sgci(ws, annee=2026):
    """Parse la feuille SGCI GL APO → liste de dicts."""
    rows = list(ws.iter_rows(values_only=True))
    records = []
    current_mois = None

    for row in rows:
        # Détecte l'en-tête de section (ex: 'JANVIER 26')
        cell3 = str(row[3] or '')
        mois = detect_month_from_header(cell3)
        if mois:
            current_mois = mois
            continue

        if current_mois is None:
            continue

        # Débit = col[4], Crédit = col[5]
        debit  = row[4]
        credit = row[5]
        libelle = (row[3] or '').strip()

        if not libelle:
            continue

        # On ne prend que les débits (sortie d'argent)
        if not isinstance(debit, (int, float)) or debit <= 0:
            continue

        montant = int(round(debit))
        cat = categorize_banque(libelle)
        if cat == '_skip':
            continue

        records.append({
            'mois':           current_mois,
            'annee':          annee,
            'banque':         'SGCI',
            'date_operation': safe_date(row[0]),
            'date_valeur':    safe_date(row[7]),
            'libelle':        libelle,
            'montant_fcfa':   montant,
            'categorie':      cat,
        })

    return records

def parse_bda(ws, annee=2026):
    """Parse la feuille BDA GL APO → liste de dicts."""
    rows = list(ws.iter_rows(values_only=True))
    records = []
    current_mois = None

    for row in rows:
        cell3 = str(row[3] or '')
        mois = detect_month_from_header(cell3)
        if mois:
            current_mois = mois
            continue

        if current_mois is None:
            continue

        debit   = row[4]
        libelle = (row[3] or '').strip()

        if not libelle:
            continue
        if not isinstance(debit, (int, float)) or debit <= 0:
            continue

        montant = int(round(debit))
        cat = categorize_banque(libelle)
        if cat == '_skip':
            continue

        # date_valeur → col[6] pour BDA
        records.append({
            'mois':           current_mois,
            'annee':          annee,
            'banque':         'BDA',
            'date_operation': safe_date(row[0]),
            'date_valeur':    safe_date(row[6]) if len(row) > 6 else None,
            'libelle':        libelle,
            'montant_fcfa':   montant,
            'categorie':      cat,
        })

    return records

# ── MAIN ─────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    if not DB_PASSWORD:
        sys.exit('❌  DB_PASSWORD manquant dans .env')

    print(f'📂  Lecture fichier : {EXCEL_PATH}')
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    records = parse_sgci(wb['SGCI GL APO']) + parse_bda(wb['BDA GL APO'])
    print(f'   → {len(records)} lignes à importer (après filtrage)')

    print('\n🔌  Connexion PostgreSQL …')
    conn = psycopg2.connect(
        host=DB_HOST, port=DB_PORT,
        user=DB_USER, password=DB_PASSWORD,
        dbname=DB_NAME, sslmode='require',
    )
    conn.autocommit = True
    cur = conn.cursor()

    # ── 1. Création de la table ───────────────────────────────────────────────
    print('\n📦  Création table banque_apo …')
    cats_sql = ", ".join(f"'{c}'" for c in ALL_CATEGORIES)
    cur.execute(f"""
    CREATE TABLE IF NOT EXISTS banque_apo (
        id             SERIAL PRIMARY KEY,
        periode_id     INTEGER NOT NULL REFERENCES periodes(id) ON DELETE CASCADE,
        banque         TEXT NOT NULL CHECK (banque IN ('SGCI','BDA')),
        date_operation DATE,
        date_valeur    DATE,
        libelle        TEXT NOT NULL,
        montant_fcfa   BIGINT NOT NULL,
        categorie      TEXT NOT NULL CHECK (categorie IN ({cats_sql}))
    );
    CREATE INDEX IF NOT EXISTS idx_banque_periode ON banque_apo(periode_id);
    CREATE INDEX IF NOT EXISTS idx_banque_cat     ON banque_apo(categorie);
    """)
    print('  ✅  Table prête.')

    # ── 2. Récupérer les periode_id ───────────────────────────────────────────
    cur.execute('SELECT id, annee, mois FROM periodes ORDER BY annee, mois')
    periodes = {(r[1], r[2]): r[0] for r in cur.fetchall()}

    # ── 3. Vidange & réinsertion ──────────────────────────────────────────────
    print('\n🗑   Vidage banque_apo …')
    cur.execute('DELETE FROM banque_apo WHERE id >= 0')

    print('💾  Insertion …')
    inserted = skipped = 0
    by_cat = {}
    by_mois = {}

    for rec in records:
        pid = periodes.get((rec['annee'], rec['mois']))
        if pid is None:
            print(f'  ⚠️  Période ({rec["annee"]}-{rec["mois"]}) introuvable — ligne ignorée : {rec["libelle"][:60]}')
            skipped += 1
            continue

        cur.execute("""
            INSERT INTO banque_apo
                (periode_id, banque, date_operation, date_valeur, libelle, montant_fcfa, categorie)
            VALUES (%s,%s,%s,%s,%s,%s,%s)
        """, (pid, rec['banque'], rec['date_operation'], rec['date_valeur'],
              rec['libelle'], rec['montant_fcfa'], rec['categorie']))
        inserted += 1

        by_cat[rec['categorie']] = by_cat.get(rec['categorie'], 0) + rec['montant_fcfa']
        key = f"{rec['annee']}-{rec['mois']:02d}"
        by_mois[key] = by_mois.get(key, 0) + rec['montant_fcfa']

    print(f'\n✅  {inserted} lignes insérées, {skipped} ignorées.\n')

    # ── 4. Rapport par catégorie ──────────────────────────────────────────────
    print('📊  Répartition par catégorie :')
    for cat in ALL_CATEGORIES:
        mt = by_cat.get(cat, 0)
        if mt:
            print(f'  {cat:<22} {mt:>15,.0f} FCFA')
    total = sum(by_cat.values())
    print(f'  {"TOTAL":<22} {total:>15,.0f} FCFA')

    print('\n📅  Répartition par mois :')
    for mois, mt in sorted(by_mois.items()):
        print(f'  {mois}  →  {mt:>15,.0f} FCFA')

    cur.close()
    conn.close()
    print('\n✅  ETL banque terminé.')
